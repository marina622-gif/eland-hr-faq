"""
이랜드건설 HR FAQ 챗봇 — Claude RAG 버전
- 데이터   : docs/ 폴더 (xlsx, pdf, txt, docx)
- 벡터 DB  : ChromaDB + DefaultEmbeddingFunction (all-MiniLM-L6-v2, ONNX)
- 생성     : Anthropic Claude API (claude-haiku-4-5-20251001)
- 관리자   : /admin  (FAQ CRUD, 미답변 질문 관리)
- 실행     : python app.py  →  http://localhost:5000
"""

import os
import sqlite3
import smtplib
import threading
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps
from pathlib import Path

import anthropic
import chromadb
import openpyxl
from chromadb.utils.embedding_functions import DefaultEmbeddingFunction
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request, session

load_dotenv()

# ── 경로 · 상수 ───────────────────────────────────────────────
BASE       = Path(__file__).parent
FAQ_FILE   = BASE / "docs" / "eland_hr_faq.xlsx"
DB_FILE    = BASE / "unanswered.db"
CHROMA_DIR = BASE / "chroma_db"
SHEET      = "FAQ"
B_COL, Q_COL, A_COL = 1, 2, 3
DATA_ROW   = 2

CAT_ORDER = ["연차/휴가", "급여/계약", "경조사", "보험/복리후생", "서류/증명", "근무"]
N_RESULTS = 5          # RAG 검색 청크 수
MAX_DIST  = 1.3        # cosine distance 상한 (초과 시 미매칭 처리)

ANTHROPIC_KEY  = os.getenv("ANTHROPIC_API_KEY", "")
CLAUDE_MODEL   = "claude-haiku-4-5-20251001"

ADMIN_EMAIL = os.getenv("ADMIN_EMAIL", "")
SMTP_HOST   = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT   = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER   = os.getenv("EMAIL_USER", "")
SMTP_PASS   = os.getenv("EMAIL_PASSWORD", "")

# Claude에게 전달하는 시스템 프롬프트
SYSTEM_PROMPT = """\
당신은 이랜드건설의 HR(인사) FAQ 챗봇입니다.

[절대 규칙]
1. 반드시 한국어로만 답변하세요.
2. 반드시 아래 [참고 문서] 내용만 근거로 답변하세요. 문서 외 지식은 사용하지 마세요.
3. [참고 문서]에서 질문과 관련된 정보를 찾을 수 없으면, 다른 말 없이 정확히 이 한 단어만 출력하세요: UNABLE_TO_ANSWER
4. 이랜드건설 HR/인사 정책(휴가, 급여, 경조사, 복리후생, 서류 등)과 무관한 질문에는 UNABLE_TO_ANSWER만 출력하세요.
5. 답변은 친절하고 간결하게, 핵심 정보 위주로 작성하세요.
6. 목록이 있으면 번호나 기호로 정리해 가독성을 높이세요.
"""

# ── ChromaDB / 임베딩 설정 ────────────────────────────────────
_collection  = None
_index_ready = False
_index_lock  = threading.Lock()


def _get_collection():
    global _collection
    if _collection is None:
        CHROMA_DIR.mkdir(exist_ok=True)
        ef = DefaultEmbeddingFunction()
        client = chromadb.PersistentClient(path=str(CHROMA_DIR))
        _collection = client.get_or_create_collection(
            name="hr_docs",
            embedding_function=ef,
            metadata={"hnsw:space": "cosine"},
        )
    return _collection


# ── 문서 청크 로드 ────────────────────────────────────────────
def _load_chunks():
    """docs/ 폴더의 모든 문서를 RAG 청크 리스트로 반환"""
    chunks = []

    # 1. Excel FAQ — Q&A 쌍을 하나의 청크로
    if FAQ_FILE.exists():
        try:
            wb = openpyxl.load_workbook(str(FAQ_FILE), data_only=True)
            ws = wb[SHEET]
            for i, row in enumerate(list(ws.iter_rows(values_only=True))[DATA_ROW:]):
                if len(row) <= max(Q_COL, A_COL):
                    continue
                cat = str(row[B_COL]).strip() if row[B_COL] else "기타"
                q   = str(row[Q_COL]).strip() if row[Q_COL] else ""
                a   = str(row[A_COL]).strip() if row[A_COL] else ""
                if q and a and cat not in ("카테고리", "None"):
                    chunks.append({
                        "id":   f"faq_{i}",
                        "text": f"질문: {q}\n답변: {a}",
                        "meta": {"source": FAQ_FILE.name, "cat": cat, "type": "faq"},
                    })
        except Exception as e:
            print(f"[RAG] Excel 로드 오류: {e}")

    # 2. TXT 파일 — 빈 줄 기준으로 단락 분리
    for fp in sorted((BASE / "docs").rglob("*.txt")):
        try:
            text = fp.read_text(encoding="utf-8", errors="ignore")
            for j, para in enumerate(
                p.strip() for p in text.split("\n\n") if len(p.strip()) > 50
            ):
                chunks.append({
                    "id":   f"txt_{fp.stem}_{j}",
                    "text": para,
                    "meta": {"source": fp.name, "type": "text"},
                })
        except Exception as e:
            print(f"[RAG] {fp.name} 로드 오류: {e}")

    # 3. PDF 파일 — 페이지 단위
    for fp in sorted((BASE / "docs").rglob("*.pdf")):
        try:
            import pdfplumber
            with pdfplumber.open(str(fp)) as pdf:
                for pg, page in enumerate(pdf.pages):
                    t = (page.extract_text() or "").strip()
                    if t:
                        chunks.append({
                            "id":   f"pdf_{fp.stem}_{pg}",
                            "text": t,
                            "meta": {"source": fp.name, "type": "pdf", "page": pg + 1},
                        })
        except ImportError:
            print("[RAG] PDF 지원: pip install pdfplumber")
        except Exception as e:
            print(f"[RAG] {fp.name} PDF 오류: {e}")

    # 4. DOCX 파일 — 단락 단위
    for fp in sorted((BASE / "docs").rglob("*.docx")):
        try:
            from docx import Document
            doc = Document(str(fp))
            for j, p in enumerate(doc.paragraphs):
                if p.text.strip() and len(p.text.strip()) > 30:
                    chunks.append({
                        "id":   f"docx_{fp.stem}_{j}",
                        "text": p.text.strip(),
                        "meta": {"source": fp.name, "type": "docx"},
                    })
        except ImportError:
            print("[RAG] DOCX 지원: pip install python-docx")
        except Exception as e:
            print(f"[RAG] {fp.name} DOCX 오류: {e}")

    return chunks


# ── 벡터 인덱스 구축 ──────────────────────────────────────────
def build_index():
    global _index_ready
    with _index_lock:
        print("[RAG] 벡터 인덱스 구축 중... (첫 실행 시 모델 다운로드로 수분 소요)")
        try:
            chunks = _load_chunks()
            if not chunks:
                print("[RAG] 인덱싱할 문서 없음")
                return

            col = _get_collection()

            # 이미 인덱스가 있으면 건너뜀 (ingest.py가 선행 실행된 경우)
            existing = col.get()
            if existing["ids"]:
                _index_ready = True
                print(f"[RAG] 기존 인덱스 사용: {len(existing['ids'])}개 청크")
                return

            # 기존 데이터 삭제 후 재구축
            existing = col.get()
            if existing["ids"]:
                col.delete(ids=existing["ids"])

            # 64개 단위 배치로 삽입 (메모리 효율)
            BATCH = 64
            for i in range(0, len(chunks), BATCH):
                b = chunks[i : i + BATCH]
                col.add(
                    ids       = [c["id"]   for c in b],
                    documents = [c["text"] for c in b],
                    metadatas = [c["meta"] for c in b],
                )

            _index_ready = True
            print(f"[RAG] 완료: {len(chunks)}개 청크 인덱싱")
        except Exception as e:
            print(f"[RAG] 인덱스 구축 실패: {e}")
            import traceback
            traceback.print_exc()


def _rebuild_bg():
    """관리자 FAQ 변경 후 비동기 인덱스 재구축"""
    threading.Thread(target=build_index, daemon=True).start()


# ── RAG 답변 생성 ─────────────────────────────────────────────
def rag_answer(question: str) -> dict:
    """ChromaDB 검색 → Claude 답변 생성"""
    if not ANTHROPIC_KEY:
        return {"error": "ANTHROPIC_API_KEY가 .env에 설정되지 않았습니다."}

    if not _index_ready:
        return {"answer": "FAQ 인덱스를 초기화 중입니다. 잠시 후 다시 질문해 주세요."}

    try:
        col     = _get_collection()
        results = col.query(
            query_texts = [question],
            n_results   = N_RESULTS,
            include     = ["documents", "distances"],
        )
        docs      = results["documents"][0]
        distances = results.get("distances", [[]])[0]

        # 모든 결과가 너무 멀면 → 담당자 연결
        if not docs or all(d > MAX_DIST for d in distances):
            return {"needs_confirm": True}

        context = "\n\n---\n\n".join(docs)

        client = anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        msg = client.messages.create(
            model      = CLAUDE_MODEL,
            max_tokens = 1000,
            system     = SYSTEM_PROMPT,
            messages   = [{
                "role":    "user",
                "content": f"[참고 문서]\n{context}\n\n[질문]\n{question}",
            }],
        )
        answer = msg.content[0].text.strip()

        if "UNABLE_TO_ANSWER" in answer:
            return {"needs_confirm": True}

        return {"answer": answer}

    except anthropic.APIError as e:
        print(f"[API] Claude 오류: {e}")
        return {"error": f"Claude API 오류가 발생했습니다. ({type(e).__name__})"}
    except Exception as e:
        print(f"[RAG] 쿼리 오류: {e}")
        return {"error": "답변 생성 중 오류가 발생했습니다."}


# ── FAQ 메모리 로드 (관리자·카테고리 탭용) ────────────────────
FAQ: list = []


def reload_faq():
    global FAQ
    if not FAQ_FILE.exists():
        FAQ = []
        return
    try:
        wb = openpyxl.load_workbook(str(FAQ_FILE), data_only=True)
        ws = wb[SHEET]
        items = []
        for row in list(ws.iter_rows(values_only=True))[DATA_ROW:]:
            if len(row) <= max(Q_COL, A_COL):
                continue
            cat = str(row[B_COL]).strip() if row[B_COL] else "기타"
            q   = str(row[Q_COL]).strip() if row[Q_COL] else ""
            a   = str(row[A_COL]).strip() if row[A_COL] else ""
            if q and a and cat not in ("카테고리", "None"):
                items.append({"cat": cat, "q": q, "a": a})
        FAQ = items
        print(f"[FAQ] {len(FAQ)}개 항목 로드")
    except Exception as e:
        print(f"[ERROR] FAQ 로드 실패: {e}")
        FAQ = []


# ── 앱 초기화 ─────────────────────────────────────────────────
reload_faq()

# ingest.py가 선행 실행된 경우 ONNX 로딩 없이 즉시 ready 처리
def _try_quick_ready():
    global _index_ready
    try:
        _c = chromadb.PersistentClient(path=str(CHROMA_DIR))
        _col = _c.get_collection(name="hr_docs")
        cnt = _col.count()
        if cnt > 0:
            _index_ready = True
            print(f"[RAG] 기존 인덱스 감지: {cnt}개 청크 — 즉시 ready")
            return True
    except Exception:
        pass
    return False

if not _try_quick_ready():
    threading.Thread(target=build_index, daemon=True).start()  # 백그라운드 인덱싱

# ── SQLite ────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(str(DB_FILE))
    conn.row_factory = sqlite3.Row
    return conn


with get_db() as _c:
    _c.execute("""
        CREATE TABLE IF NOT EXISTS unanswered (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT NOT NULL DEFAULT '',
            email       TEXT NOT NULL DEFAULT '',
            question    TEXT NOT NULL,
            created_at  TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S','now','localtime')),
            status      TEXT DEFAULT 'pending',
            answer      TEXT,
            answered_at TEXT
        )
    """)


# ── 이메일 ────────────────────────────────────────────────────
def send_email(to: str, subject: str, body: str) -> bool:
    if not all([SMTP_USER, SMTP_PASS, to]):
        return False
    try:
        msg = MIMEMultipart()
        msg["From"] = SMTP_USER
        msg["To"]   = to
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain", "utf-8"))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=10) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASS)
            s.send_message(msg)
        print(f"[MAIL] → {to}")
        return True
    except Exception as e:
        print(f"[MAIL] 실패: {e}")
        return False


# ── Flask 앱 ──────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.getenv("ADMIN_SECRET", "eland-hr-admin-2025")
ADMIN_PASS_CFG = os.getenv("ADMIN_PASS", "admin1234")


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_ok"):
            return jsonify({"error": "인증 필요", "auth": False}), 401
        return f(*args, **kwargs)
    return decorated


# ── 공개 라우트 ───────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/status")
def status():
    return jsonify({"ready": bool(FAQ), "index_ready": _index_ready})


@app.route("/faq_list")
def faq_list():
    ordered = CAT_ORDER + [
        c for c in dict.fromkeys(i["cat"] for i in FAQ) if c not in CAT_ORDER
    ]
    items = sorted(
        [{"q": i["q"], "cat": i["cat"]} for i in FAQ],
        key=lambda x: (ordered.index(x["cat"]) if x["cat"] in ordered else 99),
    )
    return jsonify({"categories": ["전체"] + ordered, "items": items})


@app.route("/chat", methods=["POST"])
def chat():
    data     = request.get_json(silent=True) or {}
    question = data.get("question", "").strip()
    if not question:
        return jsonify({"error": "질문을 입력해주세요."}), 400
    return jsonify(rag_answer(question))


@app.route("/chat/confirm", methods=["POST"])
def chat_confirm():
    data     = request.get_json(silent=True) or {}
    question = data.get("question", "").strip()
    name     = data.get("name",     "").strip()
    email    = data.get("email",    "").strip()
    if not question:
        return jsonify({"error": "질문이 없습니다."}), 400

    with get_db() as conn:
        cur  = conn.execute(
            "INSERT INTO unanswered (name, email, question) VALUES (?,?,?)",
            (name, email, question),
        )
        q_id = cur.lastrowid

    if ADMIN_EMAIL:
        send_email(
            ADMIN_EMAIL,
            f"[이랜드건설 HR FAQ] 미답변 질문 #{q_id}",
            f"미답변 질문이 접수되었습니다.\n\n"
            f"질문자: {name or '(미입력)'}\n"
            f"이메일: {email or '(미입력)'}\n"
            f"질문  : {question}\n\n"
            f"관리자 페이지: http://localhost:5000/admin",
        )
    if email:
        send_email(
            email,
            "[이랜드건설 HR FAQ] 문의 접수 완료",
            f"{name or '안녕하세요'}님,\n\n"
            f"문의하신 내용이 HR 담당자에게 전달되었습니다.\n"
            f"검토 후 답변이 등록되면 이 이메일로 알림드립니다.\n\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"질문: {question}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n\n"
            f"이랜드건설 HR 팀 드림",
        )

    return jsonify({
        "answer":  "담당자에게 전달되었습니다. 답변이 등록되면 입력하신 이메일로 알림드립니다.",
        "pending": True,
    })


# ── 관리자 라우트 ─────────────────────────────────────────────
@app.route("/admin")
def admin():
    return render_template("admin.html")


@app.route("/admin/login", methods=["POST"])
def admin_login():
    pw = (request.get_json(silent=True) or {}).get("password", "")
    if pw == ADMIN_PASS_CFG:
        session["admin_ok"] = True
        return jsonify({"ok": True})
    return jsonify({"error": "비밀번호가 틀렸습니다."}), 403


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("admin_ok", None)
    return jsonify({"ok": True})


@app.route("/admin/check-auth")
def admin_check_auth():
    return jsonify({"authenticated": bool(session.get("admin_ok"))})


@app.route("/admin/questions")
@admin_required
def admin_questions():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM unanswered ORDER BY created_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/admin/stats")
@admin_required
def admin_stats():
    with get_db() as conn:
        total   = conn.execute("SELECT COUNT(*) FROM unanswered").fetchone()[0]
        pending = conn.execute(
            "SELECT COUNT(*) FROM unanswered WHERE status='pending'"
        ).fetchone()[0]
    return jsonify({
        "total": total, "pending": pending,
        "faq_count": len(FAQ), "index_ready": _index_ready,
    })


@app.route("/admin/answer", methods=["POST"])
@admin_required
def admin_answer():
    data     = request.get_json(silent=True) or {}
    q_id     = data.get("id")
    answer   = data.get("answer",   "").strip()
    category = data.get("category", "기타").strip() or "기타"
    if not q_id or not answer:
        return jsonify({"error": "id와 답변이 필요합니다."}), 400

    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM unanswered WHERE id=?", (q_id,)
        ).fetchone()
        if not row:
            return jsonify({"error": "질문을 찾을 수 없습니다."}), 404
        question   = row["question"]
        user_email = row["email"]
        user_name  = row["name"]
        conn.execute(
            "UPDATE unanswered SET status='answered', answer=?, "
            "answered_at=strftime('%Y-%m-%d %H:%M:%S','now','localtime') WHERE id=?",
            (answer, q_id),
        )

    try:
        wb = openpyxl.load_workbook(str(FAQ_FILE))
        ws = wb[SHEET]
        ws.append([ws.max_row - DATA_ROW + 1, category, question, answer])
        wb.save(str(FAQ_FILE))
        reload_faq()
        _rebuild_bg()       # 벡터 인덱스 재구축
    except Exception as e:
        return jsonify({"error": f"Excel 저장 실패: {e}"}), 500

    if user_email:
        send_email(
            user_email,
            "[이랜드건설 HR FAQ] 문의하신 질문에 답변이 등록되었습니다",
            f"{user_name or '안녕하세요'}님,\n\n"
            f"문의하신 내용에 HR 담당자가 답변을 등록하였습니다.\n\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"질문: {question}\n\n"
            f"답변:\n{answer}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n\n"
            f"이랜드건설 HR 팀 드림",
        )

    return jsonify({"ok": True, "faq_count": len(FAQ)})


@app.route("/admin/reload", methods=["POST"])
@admin_required
def admin_reload():
    reload_faq()
    _rebuild_bg()
    return jsonify({"ok": True, "faq_count": len(FAQ), "index_ready": _index_ready})


@app.route("/admin/faq_items")
@admin_required
def admin_faq_items():
    return jsonify([
        {"idx": i, "cat": it["cat"], "q": it["q"], "a": it["a"]}
        for i, it in enumerate(FAQ)
    ])


@app.route("/admin/faq/add", methods=["POST"])
@admin_required
def admin_faq_add():
    data = request.get_json(silent=True) or {}
    cat  = data.get("cat", "기타").strip() or "기타"
    q    = data.get("q", "").strip()
    a    = data.get("a", "").strip()
    if not q or not a:
        return jsonify({"error": "질문과 답변은 필수입니다."}), 400
    try:
        wb = openpyxl.load_workbook(str(FAQ_FILE))
        ws = wb[SHEET]
        ws.append([ws.max_row - DATA_ROW + 1, cat, q, a])
        wb.save(str(FAQ_FILE))
        reload_faq()
        _rebuild_bg()
        return jsonify({"ok": True, "faq_count": len(FAQ)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/faq/update", methods=["POST"])
@admin_required
def admin_faq_update():
    data = request.get_json(silent=True) or {}
    idx  = data.get("idx")
    cat  = data.get("cat", "기타").strip() or "기타"
    q    = data.get("q",   "").strip()
    a    = data.get("a",   "").strip()
    if idx is None or not q or not a:
        return jsonify({"error": "필수 값 누락"}), 400
    try:
        wb = openpyxl.load_workbook(str(FAQ_FILE))
        ws = wb[SHEET]
        rows = list(ws.iter_rows())[DATA_ROW:]
        if idx >= len(rows):
            return jsonify({"error": "잘못된 인덱스"}), 400
        rows[idx][B_COL].value = cat
        rows[idx][Q_COL].value = q
        rows[idx][A_COL].value = a
        wb.save(str(FAQ_FILE))
        reload_faq()
        _rebuild_bg()
        return jsonify({"ok": True, "faq_count": len(FAQ)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/faq/delete", methods=["POST"])
@admin_required
def admin_faq_delete():
    data = request.get_json(silent=True) or {}
    idx  = data.get("idx")
    if idx is None:
        return jsonify({"error": "idx 필요"}), 400
    try:
        wb = openpyxl.load_workbook(str(FAQ_FILE))
        ws = wb[SHEET]
        ws.delete_rows(idx + DATA_ROW + 1)
        wb.save(str(FAQ_FILE))
        reload_faq()
        _rebuild_bg()
        return jsonify({"ok": True, "faq_count": len(FAQ)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/admin/delete_question", methods=["POST"])
@admin_required
def admin_delete_question():
    data = request.get_json(silent=True) or {}
    q_id = data.get("id")
    if not q_id:
        return jsonify({"error": "id 필요"}), 400
    with get_db() as conn:
        conn.execute("DELETE FROM unanswered WHERE id=?", (q_id,))
    return jsonify({"ok": True})


# ── 시작 ──────────────────────────────────────────────────────
if __name__ == "__main__":
    sep = "=" * 58
    print(sep)
    print("  이랜드건설 HR FAQ 챗봇  (Claude RAG)")
    print(sep)
    print(f"  FAQ 항목  : {len(FAQ)}개  ({FAQ_FILE.name})")
    print(f"  LLM 모델  : {CLAUDE_MODEL}")
    print(f"  임베딩    : DefaultEmbeddingFunction (all-MiniLM-L6-v2)")
    print(f"  챗봇      : http://localhost:5000")
    print(f"  관리자    : http://localhost:5000/admin")
    print(f"  ※ 백그라운드에서 벡터 인덱싱 중... (첫 실행 시 모델 로딩 수분 소요)")
    if not ANTHROPIC_KEY:
        print("  [WARN] ANTHROPIC_API_KEY 미설정!")
    if not SMTP_USER:
        print("  [WARN] 이메일 미설정 (EMAIL_USER, EMAIL_PASSWORD)")
    print(sep)
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
