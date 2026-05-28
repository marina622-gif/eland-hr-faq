"""
make_seed.py
docs/ 폴더의 xlsx 파일을 분석해서 FAQ_시드 시트를 자동 생성합니다.
사용법: python make_seed.py
"""

import os
import sys
from pathlib import Path

import anthropic
import openpyxl
from dotenv import load_dotenv

load_dotenv()

BASE      = Path(__file__).parent
DOCS_DIR  = BASE / "docs"
MAIN_FAQ  = "eland_hr_faq.xlsx"
API_KEY   = os.getenv("ANTHROPIC_API_KEY", "")

EXISTING_CATS = ["연차/휴가", "급여/계약", "경조사", "보험/복리후생", "서류/증명", "근무", "평가/승진"]

SEED_PROMPT = """\
당신은 HR 문서 분석 전문가입니다.
아래 엑셀 시트 내용을 분석해서 직원들이 자주 물어볼 법한 FAQ를 만들어주세요.

[규칙]
1. 질문은 직원 입장에서 실제로 궁금할 내용으로 작성
2. 답변은 시트 내용에 근거해서 구체적으로 작성
3. 카테고리는 아래 목록 중에서 선택. 해당 없으면 적절한 이름으로 새로 만들기:
   {cats}
4. 최소 5개, 최대 15개 Q&A 생성
5. 반드시 아래 형식으로만 출력 (다른 말 없이):
카테고리|질문|답변
카테고리|질문|답변
...

[시트 내용]
{content}
"""


def sheet_to_text(ws) -> str:
    lines = []
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if cells:
            lines.append(" | ".join(cells))
    return "\n".join(lines)


def generate_seed(content: str, client: anthropic.Anthropic) -> list[dict]:
    prompt = SEED_PROMPT.format(
        cats=", ".join(EXISTING_CATS),
        content=content[:6000],
    )
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2000,
        messages=[{"role": "user", "content": prompt}],
    )
    items = []
    for line in msg.content[0].text.strip().splitlines():
        parts = line.split("|")
        if len(parts) == 3:
            cat, q, a = [p.strip() for p in parts]
            if q and a:
                items.append({"cat": cat, "q": q, "a": a})
    return items


def process_file(xlsx_path: Path, client: anthropic.Anthropic):
    print(f"\n처리 중: {xlsx_path.name}")
    wb = openpyxl.load_workbook(str(xlsx_path))

    seed_sheets = [s for s in wb.sheetnames if "FAQ" in s or "시드" in s]
    data_sheets = [s for s in wb.sheetnames if s not in seed_sheets]

    if not data_sheets:
        print("  → 데이터 시트 없음, 스킵")
        return

    all_content = []
    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        text = sheet_to_text(ws)
        if text.strip():
            all_content.append(f"[시트: {sheet_name}]\n{text}")

    if not all_content:
        print("  → 내용 없음, 스킵")
        return

    combined = "\n\n".join(all_content)
    print(f"  → Claude로 FAQ 생성 중...")
    items = generate_seed(combined, client)

    if not items:
        print("  → FAQ 생성 실패")
        return

    seed_name = xlsx_path.stem + "_FAQ_시드"
    if seed_name in wb.sheetnames:
        del wb[seed_name]
    ws_seed = wb.create_sheet(seed_name)
    ws_seed.append([f"{xlsx_path.stem} FAQ 시드 (자동 생성)"])
    ws_seed.append(["카테고리", "질문", "답변"])
    for item in items:
        ws_seed.append([item["cat"], item["q"], item["a"]])

    wb.save(str(xlsx_path))
    print(f"  → 완료: {len(items)}개 FAQ 시드 생성 ({seed_name})")


def main():
    if not API_KEY:
        print("오류: ANTHROPIC_API_KEY가 설정되지 않았습니다.")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=API_KEY)
    xlsx_files = [
        p for p in DOCS_DIR.glob("*.xlsx")
        if p.name != MAIN_FAQ and not p.name.startswith("~$")
    ]

    if not xlsx_files:
        print("처리할 파일이 없습니다.")
        return

    for xlsx_path in xlsx_files:
        try:
            process_file(xlsx_path, client)
        except Exception as e:
            print(f"  → 오류: {e}")

    print("\n완료! 생성된 FAQ 시드를 GitHub에 push하면 챗봇에 자동 반영됩니다.")


if __name__ == "__main__":
    main()
